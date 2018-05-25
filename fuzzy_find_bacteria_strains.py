import xlrd
from openpyxl import Workbook
from openpyxl import load_workbook

print("Enter strain column: ")
strain_col = input()
print("Enter risk group column: ")
risk_col = input()
print("Enter risk output columnn: ")
risk_out_col = input()
print("Enter ""Match?"" output columnn: ")
match_out_col = input()
print("Enter start row number: ")
start_row = int(input())
print("Enter end row number: ")
end_row = int(input())

MasterExcelFile = "jenna_test.xlsx" # Do not change this file!
TestFile = "bacteria_test_test.xlsx" # Change this file to desired Excel workbook

workbook_master = load_workbook(MasterExcelFile)
workbook_test = load_workbook(TestFile)

MasterWorksheet = "Bacteria_Risk" # Do not change
TestWorksheet = "Strain_List" # Change to sheet in the workbook

worksheet_master = workbook_master[MasterWorksheet]
worksheet_test = workbook_test[TestWorksheet]

master_col_string = "BCDE"

found_strains = 0
no_strains = 0
risk_matches = 0
no_matches = 0


for test_row in range(start_row,end_row+1):
    print("Current row: " + str(test_row))
    current_strain = worksheet_test[strain_col.upper() + str(test_row)].value
    current_risk = worksheet_test[risk_col.upper() + str(test_row)].value
    print("Test strain: " + current_strain)
    print("Test strain risk: " + str(current_risk))

    for master_row in range(4,12546):
        master_strain = worksheet_master["A" + str(master_row)].value
        
        if master_strain is not None and master_strain == current_strain:
            found_strains = found_strains + 1
            for i in master_col_string:
                if worksheet_master[i + str(master_row)].value is not None:
                    master_risk = str(worksheet_master[i + str(master_row)].value)
                    worksheet_test[risk_out_col.upper() + str(test_row)] = master_risk
                    if master_risk[0] == str(current_risk):
                        risk_matches = risk_matches + 1
                        worksheet_test[match_out_col.upper() + str(test_row)] = "Match"
                        workbook_test.save(TestFile)
                        print("Match")
                    elif i == "E" and master_risk[0] != str(current_risk):
                        no_matches = no_matches + 1
                        worksheet_test[match_out_col.upper() + str(test_row)] = "No Match"
                        workbook_test.save(TestFile)
                        print("No Match")

        elif master_row == 12545 and worksheet_test[match_out_col.upper() + str(test_row)].value is None:
            no_strains = no_strains + 1
            worksheet_test[risk_out_col.upper() + str(test_row)] = "No strain found"
            worksheet_test[match_out_col.upper() + str(test_row)] = "No strain found"
            workbook_test.save(TestFile)
            print("No strain found")

        master_row += 1

    print("--------------------------------------------")
    if test_row == end_row:
        print("Number of strains found: " + str(found_strains))
        print("Number of strains not found: " + str(no_strains))
        print("Number of matched risks: " + str(risk_matches))
        print("Number of unmatched risks: " + str(no_matches))
    test_row += 1
    













                        
